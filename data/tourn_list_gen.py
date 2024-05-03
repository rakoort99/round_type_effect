import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from tqdm import tqdm
import argparse

parser = argparse.ArgumentParser("tourn_list_gen")
parser.add_argument(
    "tourn_list_output",
    help=".xlsx file to write list of tournaments to",
    type=str,
    nargs='?',
    const="tourn_list_final1.xlsx",
    default="tourn_list_final1.xlsx",
)
parser.add_argument(
    "start_season",
    help="first debate season to scrape tournaments from",
    type=int,
    nargs='?',
    const=2013,
    default=2013,
)
parser.add_argument(
    "end_season",
    help="last debate season to scrape tournaments from",
    type=int,
    nargs='?',
    const=2024,
    default=2024,
)

args = parser.parse_args()


def get_tourn_list(year_list, filename="tourney_list.xlsx"):
    """crawls through tabroom home page for NDT/CEDA tournaments in given years and saves list .xlsx"""
    t = 15
    # declare workbook
    wb = Workbook()

    # define webdriver, load page. we use firefox because it's best
    driver = webdriver.Firefox()
    driver.get("https://www.tabroom.com/index/index.mhtml")

    # note: it looks like this workflow makes no sense, it's not because I'm being stupid, it's
    # because the website is really badly made.
    # sleep statements are included as insurance. webdriver waits are inconsistent with javascript
    time.sleep(3)
    # select the dropdown element with circuits, send keys to select NDT/CEDA.
    circuitbox = WebDriverWait(driver, timeout=t).until(
        lambda d: d.find_elements(By.CSS_SELECTOR, ".select2-selection")
    )[0]
    time.sleep(1)
    circuitbox.click()
    time.sleep(1)

    circuitbox1 = WebDriverWait(driver, timeout=t).until(
        lambda d: d.find_elements(By.CSS_SELECTOR, ".select2-search__field")
    )[0]
    circuitbox1.send_keys("NDT")
    circuitbox1.send_keys(Keys.ENTER)
    time.sleep(5)

    # we iterate through years, puling up the list of tournaments in each year
    for year in year_list:
        success = False
        while not success:
            try:
                yearbox = WebDriverWait(driver, timeout=t).until(
                    lambda d: d.find_elements(By.CSS_SELECTOR, ".select2-selection")
                )[1]
                time.sleep(1)
                yearbox.click()
                time.sleep(1)

                yearbox1 = WebDriverWait(driver, timeout=t).until(
                    lambda d: d.find_elements(By.CSS_SELECTOR, ".select2-search__field")
                )[0]
                yearbox1.send_keys(year)
                yearbox1.send_keys(Keys.ENTER)

                time.sleep(3)

                circuitbox = WebDriverWait(driver, timeout=t).until(
                    lambda d: d.find_elements(By.CSS_SELECTOR, ".select2-selection")
                )[0]
                time.sleep(1)
                circuitbox.click()
                time.sleep(1)

                circuitbox1 = WebDriverWait(driver, timeout=t).until(
                    lambda d: d.find_elements(By.CSS_SELECTOR, ".select2-search__field")
                )[0]
                circuitbox1.send_keys("NDT")
                circuitbox1.send_keys(Keys.ENTER)
                time.sleep(3)
                success = True
            except Exception as e:
                print(e)
                print("trying again")
        # now that we have made selections, we extract tournaments from page
        tourneys = WebDriverWait(driver, timeout=t).until(
            lambda d: d.find_elements(
                By.CSS_SELECTOR, ".white.smallish.nearfull.padvertless"
            )
        )

        # create sheet for given year with list of tournaments, link to tournament pages
        sheet = wb.create_sheet(year)
        names = [x.text for x in tourneys]
        links = [x.get_attribute("href") for x in tourneys]
        for i, (name, link) in enumerate(zip(names, links)):
            sheet.cell(row=i + 1, column=1).value = name
            sheet.cell(row=i + 1, column=1).hyperlink = link

    # delete base sheet in workbook and save
    del wb["Sheet"]
    wb.save(filename)
    driver.close()


def main(output_name, start, end):
    """runs get_tourn_list"""
    print(
        f"Generating list of tournaments from season {start} to season {end}. Time is now: {datetime.now()}\n"
    )
    tqdm(get_tourn_list([str(i) for i in range(start, end + 1)], output_name))
    print(f"Done generating tournament list. Time is now: {datetime.now()}\n")


if __name__ == "__main__":
    main(args.tourn_list_output, args.start_season, args.end_season)
